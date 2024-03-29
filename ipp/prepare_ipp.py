import pandas as pd
import glob
import os
import numpy as np
from sklearn import preprocessing
import openpyxl
import codecs
from openpyxl.utils import get_column_letter

ws = ["w2", "w3"]
w2_remove_cols = ["HoleArea", "HoleRatio", "Holes", "Stats", "Density(red)", "Density(gree)", "Density(blue)",
                  "Segmentation"]
w3_cols = ['Density(mean)', 'Density(min)', 'Density(max)', 'Density(std.)', 'Density(sum)']
good_cols = ["Roundness", "Aspect", "Density(mean)"]


def txt_to_excel(filename, outfile):
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 0
    lines = codecs.open(filename, 'r')
    for line in lines:
        row += 1
        line = line.strip()
        line = line.split('\t')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].strip().format(get_column_letter(col)))
    wb.save(outfile)


def remove_cols(all, remove):
    keep = [a for a in all if a not in remove]
    return keep


def extract_data(folder):
    for w in ws:
        path = f"{folder}/{w}.xlsx"
        print(path)
        txt_path = path.replace(".xlsx", ".xls")
        if not os.path.exists(txt_path):
            continue
        txt_to_excel(txt_path, path)
        extract_excel(path)


def extract_excel(path):
    df = pd.read_excel(path)
    df = df.applymap((lambda x: "".join(x.split()) if type(x) is str else x))
    columns = df.loc[1].tolist()
    df.columns = columns
    w2_cols = remove_cols(columns, w2_remove_cols)
    stats = df['Stats'].tolist()
    mean_index = [i for i, x in enumerate(stats) if x == 'Mean']
    if "w1" in path:
        new_df = df.iloc[mean_index, :]
    else:
        means = []
        for i in range(0, len(mean_index), 2):
            correct_mean = df.iloc[mean_index[i + 1], 1]
            df.iloc[mean_index[i]].loc['Roundness'] = correct_mean
            if "w3" in path:
                means.append(df.loc[mean_index[i], w3_cols].tolist())
            else:
                means.append(df.loc[mean_index[i], w2_cols].tolist())
        if "w3" in path:
            new_df = pd.DataFrame(np.array(means), columns=w3_cols)
        else:
            new_df = pd.DataFrame(np.array(means), columns=w2_cols)
    new_path = os.path.splitext(path)[0] + '_new.csv'
    new_df.to_csv(new_path, index=None)


def concat_data(folder, lib):
    name = os.path.split(folder)[1]

    dfs = []
    for w in ws:
        path = f"{folder}/{w}_new.csv"
        df = pd.read_csv(path, low_memory=False)
        dfs.append(df)
    new_df = pd.concat(dfs, axis=1)

    ts = []
    for i in range(0, len(new_df), 16):
        ts += range(1, 17)
    new_df["label"] = name
    new_df["time"] = ts
    new_df["lib"] = lib
    new_df["filename"] = os.listdir(folder + "/w1")
    new_df.to_csv(f"{folder}/merge.csv", index=None)


def compute_zscore(path, save_path):
    df = pd.read_csv(path, low_memory=False)
    values = df.values[:, :-4].astype('float32')
    data = preprocessing.scale(values)
    df.iloc[:, :-4] = data
    df.to_csv(save_path, index=None)


def compute_zscore_control(path, save_path):
    df = pd.read_csv(path, low_memory=False)
    ids = df["label"]

    controls = []
    for i in range(0, len(df), 16):
        id = ids[i]
        if 'dmso' in id:
            values = df.iloc[i:i + 16, :-3].values
            controls.append(values)
    mean = np.mean(controls, 0)
    std = np.std(controls, 0)

    for i in range(0, len(df), 16):
        values = df.iloc[i:i + 16, :-3].values
        df.iloc[i:i + 16, :-3] = (values - mean) / std
    df.to_csv(save_path, index=None)


def fix_outlier(data):
    row, col = data.shape

    outliers = []
    for r in range(row):
        diffs = []
        for c in range(col):
            if r == 0 or r == row - 1:
                continue
            else:
                diff = round(abs(data[r, c] - (data[r - 1, c] + data[r + 1, c]) / 2), 2)
                diffs.append(diff)
        if len(diffs) > 0:
            diffs = np.array(diffs)
            num_outlier = np.sum(diffs > 2)
            if num_outlier > 5:
                outliers.append(r)
            # print(r, np.sum(diffs>2), np.max(diffs))
    if len(outliers) > 1:
        return []
    elif len(outliers) == 1:
        # for c in range(col):
        #     plt.plot(range(16), data[:, c])
        # plt.show()
        data[outliers[0], :] = (data[outliers[0] - 1, :] + data[outliers[0] + 1, :]) / 2
        # for c in range(col):
        #     plt.plot(range(16), data[:, c])
        # plt.show()
    return data


def prepare_data(path, is_full=True, rm_outlier=True):
    df = pd.read_csv(path, low_memory=False)
    ids = df["label"]
    lib = df["lib"]
    filename = df["filename"]
    if is_full:
        df = df.drop(["label", "time", "lib", "filename"], axis=1)
        save_path = path.replace(".csv", f"_stack.csv")
    else:
        df = df[good_cols]
        save_path = path.replace(".csv", f"_stack_good.csv")
    x = []
    y = []
    libs = []
    filenames = []

    for i in range(0, len(df), 16):
        id = str(ids[i])
        id = id.replace(".0", "")
        data = df.loc[i:i + 15].values.astype(float)
        if rm_outlier:
            data = fix_outlier(data)
        if len(data) > 0:
            data = data.T.reshape(-1)
            x.append(data)
            y.append(id)
            libs.append(lib[i])
            filenames.append("_".join(filename[i].split("_")[:-1]))
    new_df = pd.DataFrame(x)
    new_df["label"] = y
    new_df["lib"] = libs
    new_df["filename"] = filenames
    new_df.to_csv(save_path, index=None)


def merge_data(paths, save_path):
    dfs = []
    for path in paths:
        if os.path.exists(path):
            df = pd.read_csv(path, low_memory=False)
            dfs.append(df)
    new_df = pd.concat(dfs)
    new_df.to_csv(save_path, index=None)


def separate_lib(path, libs, save_paths):
    df = pd.read_csv(path, low_memory=False)
    lib = df["lib"].values
    for l, save_path in zip(libs, save_paths):
        index = np.where(lib == l)[0]
        new_df = df.loc[index, :]
        new_df.to_csv(save_path, index=None)


def relabel(path):
    target_df = pd.read_excel("../utils/annotation.xlsx", sheet_name="FDA-list")[["index", "single"]]
    target_df["index"] = target_df["index"].astype(str)

    df = pd.read_csv(path, low_memory=False)
    df["label"] = df["label"].astype(str)
    df = pd.merge(df, target_df, left_on="label", right_on="index", how="left")
    df["label"] = df["single"]
    del df["index"]
    del df["single"]
    df.to_csv(path, index=None)


def group_feature(path):
    df = pd.read_csv(path, low_memory=False)
    labels = df["label"].values
    del df["label"]
    del df['lib']
    filename = df["filename"].values
    for i, name in enumerate(filename):
        if "_" in name:
            name = name.split("_")[0]
        if labels[i] == 'dmso':
            name = 'dmso'
        # filename[i] = name.split("-")[0].split("^")[0]
        filename[i] = name
    df["filename"] = filename

    group_df = df.groupby("filename")
    mean_df = group_df.mean()
    # idx = mean_df.index.tolist()
    # if 'dmso' in idx:
    #     mean_df = mean_df.drop('dmso')
    if "lib" in mean_df.columns.tolist():
        del mean_df["lib"]

    mean_df.to_csv(path.replace(".csv", "_mean.csv"))


if __name__ == "__main__":
    # ls = [
    #     "l1",
    #     "l2",
    #     "l3"
    # ]
    dataset = "center"
    # merge_paths = []
    #
    # for l in ls:
    #     folders = glob.glob(f"H:/Extract features/{l}/{dataset}/*")
    #     for folder in folders:
    #         try:
    #             if not os.path.exists(folder+"/w2.xls"):
    #                 continue
    #             print(folder)
    #             extract_data(folder)
    #             concat_data(folder, l)
    #         except Exception as e:
    #             print(e)
    #     paths = [path+"/merge.csv" for path in folders]
    #     merge_path = f"H:/Extract features/ipp/{l}_{dataset}.csv"
    #     merge_paths.append(merge_path)
    #     merge_data(paths, merge_path)
    #
    merge_path = f"../paper/data/l123_{dataset}.csv"
    # merge_data(merge_paths, merge_path)
    zscore_path = merge_path.replace(".csv", "_zscore.csv")
    # compute_zscore(merge_path, zscore_path)
    # prepare_data(zscore_path, rm_outlier=True)
    # group_feature(zscore_path.replace(".csv", "_stack.csv"))
    prepare_data(merge_path, rm_outlier=False)
    group_feature(merge_path.replace(".csv", "_stack.csv"))

    # id = "290"
    # # folders = glob.glob(f"H:/Extract features/new/{id}/*")
    # # for folder in folders:
    # #     if not os.path.exists(folder+"/w2.xls"):
    # #         continue
    # #     print(folder)
    # #     extract_data(folder)
    # #     concat_data(folder, "")
    # # paths = [path+"/merge.csv" for path in folders]
    # merge_path = f"H:/Extract features/new/{id}.csv"
    # # merge_data(paths, merge_path)
    # # #relabel(merge_path)
    # zscore_path = merge_path.replace(".csv", "_zscore.csv")
    # # compute_zscore(merge_path, zscore_path)
    # prepare_data(zscore_path, rm_outlier=True)
    # group_feature(zscore_path.replace(".csv", "_stack.csv"))

